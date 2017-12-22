import time
import webbrowser
import tkinter as tk
from openpyxl import load_workbook
import random


wb2 = load_workbook('C:\MoneyJobKeeper.xlsx', data_only=True)
# finds a workbook at C:\MoneyJobKeeper.xlsx
ws2 = wb2.active  # sets only sheet on that workbook to current sheet
totalQuestions = ws2['J3']
# detects file that reads out the amount of questions in the worksheet
rando = random.randint(2, int(totalQuestions.value)+1)
# finds a random number that is assigned to a particular question
global sec


def questionDisplay():
    question = ws2['C' + str(rando)]
    # finds location of question that has been chosen
    print("\n" + question.value)  # puts out all info for the reader to see


def questionAnswer():
    wb2 = load_workbook('C:\MoneyJobKeeper.xlsx', data_only=True)
    # finds a workbook at C:\MoneyJobKeeper.xlsx
    ws2 = wb2.active  # sets only sheet on that workbook to current sheet

    name = ws2['B' + str(rando)]  # finds location of name that has been chosen
    book = ws2['D' + str(rando)]  # finds location of book that has been chosen
    pageNumber = ws2['E' + str(rando)]
    # finds location of pageNumber that has been chosen
    print("The correct answer is: " + book.value + '\n' + name.value +
          " submitted this question." + "\nThe quote can be found on page " +
          str(int(pageNumber.value)) + ".")


global pointTotal
pointTotal = 0  # running point total for round


"""
timer is going to be part of scoreboard, counting down and allowing for
you to choose either
a rebound or regular question based on user input, which will eventually
be controlled by a button
"""


def timer():

    # Establish initial variables
    global timerDone  # this will allow for the script to work multiple times
    timerDone = False
    regTime = 20
    reboundTime = 10
    global questionTypePublic
    # this will keep the type of question answered for later use
    # Timer script, counts down in variable "sec"

    def counter(sec):
            while sec != 0:
                    print(sec)
                    time.sleep(1)
                    sec = sec-1

    # Filler for clicking the choice of what kind of time someone will be using
    questionDisplay()
    timeChoice = input('regular/rebound ')
    while True:
        if(timeChoice == 'regular'):
            sec = regTime
            questionTypePublic = 'reg'
            break
        else:
            if (timeChoice == 'rebound'):
                sec = reboundTime
                questionTypePublic = 'rebound'
                break
            else:
                    print("I don't understand, can you say that again?")
    counter(sec)
    questionAnswer()
    timerDone = True


# scoreboard will add points in 2s or 3s, tallying up
# and then allowing for a reset at end of round

def scoreboard():
    while timerDone is True:
        # this checks type of points rewarded (reg/rb) and determines
        # how much a correct answer is worth
        if (questionTypePublic == 'reg'):
            roundPointsRewarded = 3
        else:
            roundPointsRewarded = 2

        # this checks if the answer is correct/incorrect
        pointsRewarded = input('Answered correctly? (y/n) ')
        while True:
            if (pointsRewarded == 'y'):
                global pointTotal
                pointTotal = pointTotal + roundPointsRewarded
                print("Nice Job! Your team gets " + str(roundPointsRewarded) +
                      " points! You now have " + str(pointTotal) + " points!")
                timer()
                break
            elif (pointsRewarded == 'n'):
                print("Better luck next time?")
                timer()
                break
            else:
                print("Say that again?")


timer()
scoreboard()
